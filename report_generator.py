# ============================================
# report_generator.py - Generate Excel Reports
# ============================================

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from config import REPORTS_FOLDER

class ReportGenerator:
    """Generates professional Excel and CSV reports"""

    def __init__(self):
        """Create output directories if they don't exist"""
        os.makedirs(REPORTS_FOLDER, exist_ok=True)
        os.makedirs("results/logs", exist_ok=True)

    # ==========================================
    # GENERATE COMPLETE EXCEL REPORT
    # ==========================================
    def generate_excel_report(self, all_results):
        """
        Create a comprehensive Excel workbook with:
        - Summary sheet
        - One sheet per niche
        - Color coding by opportunity score
        - Charts and visualizations
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{REPORTS_FOLDER}/YouTube_Niche_Analysis_{timestamp}.xlsx"
        
        print(f"\n📊 Generating Excel Report...")
        
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # ----------------------------------------
        # SHEET 1: EXECUTIVE SUMMARY
        # ----------------------------------------
        summary_ws = wb.create_sheet("📊 SUMMARY", 0)
        self._create_summary_sheet(summary_ws, all_results)
        
        # ----------------------------------------
        # SHEET 2: ALL CHANNELS COMBINED
        # ----------------------------------------
        all_ws = wb.create_sheet("🔍 ALL CHANNELS", 1)
        all_channels = []
        for niche, channels in all_results.items():
            for ch in channels:
                ch["niche"] = niche
                all_channels.append(ch)
        
        if all_channels:
            self._create_channels_sheet(all_ws, all_channels, "ALL NICHES")

        # ----------------------------------------
        # SHEET 3+: ONE SHEET PER NICHE
        # ----------------------------------------
        niche_colors = {
            "Finance & Investing": "1B5E20",
            "Business & Entrepreneurship": "0D47A1", 
            "Technology & Software": "4A148C",
            "Health & Wellness": "880E4F",
            "True Crime": "3E2723",
            "Motivational & Self Help": "E65100",
            "AI & Future Technology": "006064",
            "History & Education": "BF360C",
        }
        
        for i, (niche_name, channels) in enumerate(all_results.items()):
            if channels:
                sheet_name = niche_name[:28]  # Excel 31 char limit
                ws = wb.create_sheet(f"💰 {sheet_name}", i + 2)
                color = niche_colors.get(niche_name, "1565C0")
                self._create_channels_sheet(ws, channels, niche_name, color)

        # Save workbook
        wb.save(filename)
        print(f"✅ Excel Report Saved: {filename}")
        return filename

    # ==========================================
    # CREATE SUMMARY SHEET
    # ==========================================
    def _create_summary_sheet(self, ws, all_results):
        """Create executive summary sheet"""
        
        # Title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = "🎯 YOUTUBE NICHE OPPORTUNITY ANALYSIS - COMPLETE SUMMARY"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="1565C0")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Subtitle with timestamp
        ws.merge_cells('A2:J2')
        subtitle = ws['A2']
        subtitle.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        subtitle.font = Font(size=11, color="666666")
        subtitle.alignment = Alignment(horizontal="center")

        # Column headers for summary table
        headers = [
            "Niche", "Total Channels Found", "Golden Opportunities",
            "Best Channel", "Best Score", "Avg View/Sub Ratio",
            "CPM Min", "CPM Max", "Est. Monthly Revenue", "Recommendation"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor="2E7D32")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = self._thin_border()
        
        ws.row_dimensions[4].height = 30

        # Data rows
        row = 5
        for niche_name, channels in all_results.items():
            if not channels:
                continue
            
            # Calculate summary stats
            golden = [c for c in channels if c.get("opportunity_score", 0) >= 65]
            best = max(channels, key=lambda x: x.get("opportunity_score", 0)) if channels else None
            avg_ratio = sum(c.get("view_sub_ratio", 0) for c in channels) / len(channels) if channels else 0
            
            from niche_keywords import NICHE_KEYWORDS
            niche_data = NICHE_KEYWORDS.get(niche_name, {})
            cpm_min = niche_data.get("cpm_min", 0)
            cpm_max = niche_data.get("cpm_max", 0)
            
            # Estimate revenue
            if best:
                est_revenue = f"${cpm_min * 500:,} - ${cpm_max * 500:,}"
            else:
                est_revenue = "N/A"
            
            # Recommendation
            if len(golden) >= 5:
                rec = "🔥 HIGHLY RECOMMENDED"
                row_color = "E8F5E9"
            elif len(golden) >= 2:
                rec = "✅ RECOMMENDED"
                row_color = "F1F8E9"
            elif len(channels) >= 5:
                rec = "⚠️  CONSIDER"
                row_color = "FFF9C4"
            else:
                rec = "❌ SKIP"
                row_color = "FFEBEE"

            row_data = [
                niche_name,
                len(channels),
                len(golden),
                best.get("channel_name", "N/A") if best else "N/A",
                best.get("opportunity_score", 0) if best else 0,
                round(avg_ratio, 1),
                f"${cpm_min}",
                f"${cpm_max}",
                est_revenue,
                rec
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = PatternFill("solid", fgColor=row_color)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = self._thin_border()
                if col == 1:
                    cell.font = Font(bold=True)
            
            row += 1

        # Column widths
        column_widths = [30, 18, 18, 25, 12, 18, 10, 10, 22, 22]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    # ==========================================
    # CREATE CHANNELS DETAIL SHEET
    # ==========================================
    def _create_channels_sheet(self, ws, channels, niche_name, color="1565C0"):
        """Create detailed channel analysis sheet"""
        
        # Title row
        ws.merge_cells('A1:N1')
        title = ws['A1']
        title.value = f"💰 {niche_name.upper()} - CHANNEL OPPORTUNITIES"
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor=color)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Column headers
        headers = [
            "Grade", "Score", "Channel Name", "Channel URL",
            "Subscribers", "Total Views", "Videos",
            "View/Sub Ratio", "Views/Video", "Daily Views",
            "Channel Age", "Country", "Niche", "Action"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="37474F")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = self._thin_border()
        
        ws.row_dimensions[2].height = 25

        # Sort channels by opportunity score (highest first)
        sorted_channels = sorted(
            channels,
            key=lambda x: x.get("opportunity_score", 0),
            reverse=True
        )

        # Data rows
        for row_idx, channel in enumerate(sorted_channels, 3):
            score = channel.get("opportunity_score", 0)
            
            # Color code by score
            if score >= 80:
                row_color = "FFD700"  # Gold
            elif score >= 65:
                row_color = "C8E6C9"  # Light green
            elif score >= 50:
                row_color = "E3F2FD"  # Light blue
            elif score >= 35:
                row_color = "FFF9C4"  # Light yellow
            else:
                row_color = "FFEBEE"  # Light red

            row_data = [
                channel.get("grade_emoji", "⭐"),
                score,
                channel.get("channel_name", "Unknown"),
                channel.get("channel_url", ""),
                channel.get("subscribers", 0),
                channel.get("total_views", 0),
                channel.get("video_count", 0),
                channel.get("view_sub_ratio", 0),
                channel.get("views_per_video", 0),
                channel.get("daily_views", 0),
                f"{channel.get('channel_age_days', 0)} days",
                channel.get("country", "Unknown"),
                channel.get("niche", niche_name),
                "🎯 Study This Channel" if score >= 65 else "📊 Reference",
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = PatternFill("solid", fgColor=row_color)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = self._thin_border()
                
                # Make channel URL clickable
                if col == 4 and value:
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")
                
                # Bold channel name
                if col == 3:
                    cell.font = Font(bold=True)

        # Set column widths
        widths = [8, 8, 30, 40, 15, 18, 10, 15, 15, 15, 12, 10, 25, 22]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    # ==========================================
    # GENERATE CSV REPORT
    # ==========================================
    def generate_csv_report(self, all_results):
        """Generate simple CSV file for all channels"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{REPORTS_FOLDER}/channels_{timestamp}.csv"
        
        all_rows = []
        for niche, channels in all_results.items():
            for ch in channels:
                all_rows.append({
                    "niche": niche,
                    "grade": ch.get("grade", ""),
                    "score": ch.get("opportunity_score", 0),
                    "channel_name": ch.get("channel_name", ""),
                    "channel_url": ch.get("channel_url", ""),
                    "subscribers": ch.get("subscribers", 0),
                    "total_views": ch.get("total_views", 0),
                    "video_count": ch.get("video_count", 0),
                    "view_sub_ratio": ch.get("view_sub_ratio", 0),
                    "views_per_video": ch.get("views_per_video", 0),
                    "country": ch.get("country", ""),
                    "created_date": ch.get("created_date", ""),
                })
        
        if all_rows:
            df = pd.DataFrame(all_rows)
            df.to_csv(filename, index=False)
            print(f"✅ CSV Report Saved: {filename}")
        
        return filename

    def _thin_border(self):
        """Return thin border style for cells"""
        thin = Side(style='thin', color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)